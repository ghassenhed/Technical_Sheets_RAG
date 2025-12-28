import math
from operator import itemgetter
import csv
import pdfplumber
from PIL import ImageDraw, ImageFont, Image
from pdfplumber.table import TableFinder
import os
from pathlib import Path
from .DataSheetParsers.DataSheet import *
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def almost_equals(num1, num2, precision=5.0):
    return abs(num1 - num2) < precision


class Point:
    r = 4
    hr = r / 2
    tail = 5

    def __init__(self, *xy):
        if len(xy) == 1:
            xy = xy[0]
        self.x, self.y = xy
        self.x = math.ceil(self.x)
        self.y = math.ceil(self.y)
        self.down = False
        self.up = False
        self.left = False
        self.right = False

    @property
    def symbol(self):
        direction_table = {
            (False, False, False, False): '◦',

            (True, False, False, False): '↑',
            (False, True, False, False): '↓',
            (True, True, False, False): '↕',

            (True, True, True, False): '⊢',
            (True, True, False, True): '⊣',

            (False, False, True, False): '→',
            (False, False, False, True): '←',
            (False, False, True, True): '↔',

            (True, False, True, True): '⊥',
            (False, True, True, True): '⊤',

            (True, True, True, True): '╋',

            (True, False, True, False): '┗',
            (True, False, False, True): '┛',

            (False, True, True, False): '┏',
            (False, True, False, True): '┛',

        }
        return direction_table[(self.up, self.down, self.right, self.left)]

    def __repr__(self):
        return "Point<X:{} Y:{}>".format(self.x, self.y)

    def distance(self, other: 'Point'):
        return math.sqrt(((self.x - other.x) ** 2) + ((self.y - other.y) ** 2))

    @property
    def as_tuple(self):
        return self.x, self.y

    def draw(self, canvas: ImageDraw.ImageDraw, color='red'):
        canvas.ellipse((self.x - self.hr, self.y - self.hr, self.x + self.hr, self.y + self.hr), fill=color)
        if self.down:
            canvas.line(((self.x, self.y), (self.x, self.y + self.tail)), 'blue')
        if self.up:
            canvas.line(((self.x, self.y), (self.x, self.y - self.tail)), 'blue')
        if self.left:
            canvas.line(((self.x, self.y), (self.x - self.tail, self.y)), 'blue')
        if self.right:
            canvas.line(((self.x, self.y), (self.x + self.tail, self.y)), 'blue')

    def points_to_right(self, other_points: List['Point']):
        sorted_other_points = sorted(other_points, key=lambda other: other.x)
        filtered_other_points = filter(lambda o: almost_equals(o.y, self.y) and o != self and o.x > self.x,
                                       sorted_other_points)
        return list(filtered_other_points)

    def points_below(self, other_points: List['Point']):
        sorted_other_points = sorted(other_points, key=lambda other: other.y)
        filtered_other_points = filter(lambda o: almost_equals(o.x, self.x) and o != self and o.y > self.y,
                                       sorted_other_points)
        return list(filtered_other_points)

    def on_same_line(self, other: 'Point'):
        if self == other:
            return False
        if almost_equals(self.x, other.x) or almost_equals(self.y, other.y):
            return True
        return False

    def is_above(self, other: 'Point'):
        return self.y < other.y

    def is_to_right(self, other: 'Point'):
        return self.x > other.x

    def is_below(self, other: 'Point'):
        return self.y > other.y

    def is_to_left(self, other: 'Point'):
        return self.x < other.x

    def get_right(self, others: List['Point']):
        others = self.points_to_right(others)
        for point in others:
            if point.down:
                return point
        return None

    def get_bottom(self, others: List['Point'], left=False, right=False):
        others = self.points_below(others)
        for point in others:
            if point.up:
                if left:
                    if not point.right:
                        continue
                if right:
                    if not point.left:
                        continue
                return point
        return None

    def has_above(self, others: List['Point']):
        others = list(filter(lambda p: p.up, others))
        point = list(sorted(others, key=lambda p: p.y))[0]
        if point.is_above(self) and point.up:
            return True
        return False

    def copy(self, other: 'Point'):
        self.down = other.down
        self.up = other.up
        self.left = other.left
        self.right = other.right

    def merge(self, other: 'Point'):
        self.up |= other.up
        self.down |= other.down
        self.left |= other.left
        self.right |= other.right

    def __eq__(self, other: 'Point'):
        if not almost_equals(self.x, other.x):
            return False
        return almost_equals(self.y, other.y)

    def __hash__(self):
        return hash((self.x, self.y))


class Line:

    def __init__(self, p1: 'Point', p2: 'Point'):
        self.p1 = p1
        self.p2 = p2
        self.vertical = almost_equals(self.x, self.cx)
        if self.vertical:
            if self.p1.is_above(self.p2):
                pass
            else:
                self.p1, self.p2 = self.p2, self.p1
        else:
            if self.p2.is_to_right(self.p1):
                pass
            else:
                self.p1, self.p2 = self.p2, self.p1

        if self.vertical:
            self.p1.down = True
            self.p2.up = True
        else:
            self.p1.right = True
            self.p2.left = True

    def __hash__(self):
        return hash((self.p1, self.p2, self.vertical))

    @property
    def x(self):
        return self.p1.x

    @property
    def y(self):
        return self.p1.y

    @property
    def cx(self):
        return self.p2.x

    @property
    def cy(self):
        return self.p2.y

    @property
    def length(self):
        return self.p1.distance(self.p2)

    def __repr__(self):
        return 'Line<p1:{} p2:{} {}>'.format(self.p1, self.p2, 'vertical' if self.vertical else 'horizontal')

    def draw(self, canvas: ImageDraw.ImageDraw, color='blue'):
        x, y = self.x, self.y
        cx, cy = self.cx, self.cy

        canvas.line(((x, y), (cx, cy)), color, width=2)

    @property
    def as_tuple(self):
        return (self.x, self.y), (self.cx, self.cy)

    def infite_intersect(self, other: 'Line'):
        line1 = self.as_tuple
        line2 = other.as_tuple
        x_diff = (line1[0][0] - line1[1][0], line2[0][0] - line2[1][0])
        y_diff = (line1[0][1] - line1[1][1], line2[0][1] - line2[1][1])  # Typo was here

        def det(point_a, point_b):
            return point_a[0] * point_b[1] - point_a[1] * point_b[0]

        div = det(x_diff, y_diff)
        if div == 0:
            return None, None
        d = (det(*line1), det(*line2))
        x = det(d, x_diff) / div
        y = det(d, y_diff) / div
        return x, y

    def intersect(self, other: 'Line', print_fulness=False) -> bool:
        """ this returns the intersection of Line(pt1,pt2) and Line(ptA,ptB)
              returns a tuple: (xi, yi, valid, r, s), where
              (xi, yi) is the intersection
              r is the scalar multiple such that (xi,yi) = pt1 + r*(pt2-pt1)
              s is the scalar multiple such that (xi,yi) = pt1 + s*(ptB-ptA)
                  valid == 0 if there are 0 or inf. intersections (invalid)
                  valid == 1 if it has a unique intersection ON the segment    """
        point_1 = self.x, self.y
        point_2 = self.cx, self.cy
        point_a = other.x, other.y
        point_b = other.cx, other.cy
        if self.vertical:
            if self.y > self.cy:
                if self.y >= other.y >= self.cy:
                    pass
                else:
                    return False
        else:
            if other.y > other.cy:
                if other.y >= self.y >= other.cy:
                    pass
                else:
                    return False
        det_tolerance = 0.0001
        x1, y1 = point_1
        x2, y2 = point_2
        dx1 = x2 - x1
        dy1 = y2 - y1
        x, y = point_a
        xb, yb = point_b
        dx = xb - x
        dy = yb - y
        det = (-dx1 * dy + dy1 * dx)
        if math.fabs(det) < det_tolerance:
            return False
        det_inv = 1.0 / det
        r = det_inv * (-dy * (x - x1) + dx * (y - y1))
        s = det_inv * (-dy1 * (x - x1) + dx1 * (y - y1))
        if print_fulness:
            print('self segment', r)
            print('other segment', s)
        if r > 1 or s > 1:  # can't be higher than 1, 1 means they are NOT intersecting
            return False
        if r > -0.1 and s > -0.1:  # This can happen on edges, so we allow small inaccuracy
            return True
        return False

    def intersection(self, other: 'Line', print_fulness=False) -> (int, int):
        """ this returns the intersection of Line(pt1,pt2) and Line(ptA,ptB)
                      returns a tuple: (xi, yi, valid, r, s), where
                      (xi, yi) is the intersection
                      r is the scalar multiple such that (xi,yi) = pt1 + r*(pt2-pt1)
                      s is the scalar multiple such that (xi,yi) = pt1 + s*(ptB-ptA)
                          valid == 0 if there are 0 or inf. intersections (invalid)
                          valid == 1 if it has a unique intersection ON the segment    """
        point_1 = self.x, self.y
        point_2 = self.cx, self.cy
        point_a = other.x, other.y
        point_b = other.cx, other.cy
        det_tolerance = 1
        x1, y1 = point_1
        x2, y2 = point_2
        dx1 = x2 - x1
        dy1 = y2 - y1
        x, y = point_a
        xb, yb = point_b
        dx = xb - x
        dy = yb - y
        det = (-dx1 * dy + dy1 * dx)

        if math.fabs(det) < det_tolerance:
            return None, None
        det_inv = 1.0 / det
        r = det_inv * (-dy * (x - x1) + dx * (y - y1))
        s = det_inv * (-dy1 * (x - x1) + dx1 * (y - y1))
        xi = (x1 + r * dx1 + x + s * dx) / 2.0
        yi = (y1 + r * dy1 + y + s * dy) / 2.0
        if print_fulness:
            print('self segment', r)
            print('other segment', s)
        return (round(xi), round(yi)), round(r, 4), round(s, 4)

    def is_between(self, point: 'Point'):
        pt1 = self.p1
        pt2 = self.p2
        cross_product = (point.y - pt1.y) * (pt2.x - pt1.x) - (point.x - pt1.x) * (pt2.y - pt1.y)

        # compare versus epsilon for floating point values, or != 0 if using integers
        if abs(cross_product) > math.e:
            return False

        dot_product = (point.x - pt1.x) * (pt2.x - pt1.x) + (point.y - pt1.y) * (pt2.y - pt1.y)
        if dot_product < 0:
            return False

        squared_length_ba = (pt2.x - pt1.x) * (pt2.x - pt1.x) + (pt2.y - pt1.y) * (pt2.y - pt1.y)
        if dot_product > squared_length_ba:
            return False

        return True

    def on_line(self, point: 'Point'):
        if self.vertical:
            if almost_equals(self.p1.x, point.x):
                return True
        else:
            if almost_equals(self.p1.y, point.y):
                return True
        return False

    def __contains__(self, other: {'Line', 'Point'}):
        if type(other) == Line:
            if self.vertical == other.vertical:
                return False
            return self.intersect(other)
        if type(other) == Point:
            return self.is_between(other)
            pass

    def on_same_line(self, other: 'Line'):
        if other.vertical != self.vertical:
            return False
        if self.vertical:
            return self.x == other.x
        else:
            return self.y == other.y

    def __eq__(self, other: 'Line'):
        return self.on_same_line(other)

    def corner(self, other: 'Line'):
        if self.p1 == other.p1 or self.p2 == other.p2 or self.p1 == other.p2:
            return True
        return False

    def connected(self, other: 'Line'):
        return other.p1 in self or other.p2 in self

    def parallel(self, other: 'Line'):
        return self.vertical == other.vertical

    def on_corners(self, other: 'Point'):
        return other == self.p1 or other == self.p2

    def test_intersection(self, other: 'Line'):
        """ prints out a test for checking by hand... """
        print('Testing intersection of:')
        print('\t', self)
        print('\t', other)
        result = self.intersection(other, True)
        print("\t Intersection result =", Point(result[0]))
        print()


class Cell:
    """P1-------P2
        |       |
        |       |
        |       |
        |       |
       P4-------P3
    """
    try:
        font = ImageFont.truetype('arial', size=9)
    except:
        font = ImageFont.load_default()

    def __init__(self, p1, p2, p3, p4):
        self.p1: Point = p1
        self.p2: Point = p2
        self.p3: Point = p3
        self.p4: Point = p4
        self.text = ''
        self.words = []  # type: List[str]

    def __repr__(self):
        return 'Cell <"{}"> '.format(self.text.replace('\n', ' '))

    def get_text(self):
        return ''.join(map(itemgetter('text'),self.words))

    @property
    def clean_text(self) -> str:
        return self.text.replace('\n', ' ')

    def __hash__(self):
        return hash(self.text) + hash(self.as_tuple)

    def on_same_line(self, other: 'Cell'):
        return self.p1.on_same_line(other.p1)

    def on_same_row(self, other: 'Cell'):
        return self.p1.y == other.p1.y

    @property
    def as_tuple(self):
        return self.p1.as_tuple, self.p2.as_tuple, self.p3.as_tuple, self.p4.as_tuple

    def __eq__(self, other: 'Cell'):
        if self.p1 == other.p1 and self.p2 == other.p2 and self.p3 == other.p3 and self.p4 == other.p4:
            return True
        if self.p1 == other.p2 and self.p2 == other.p3 and self.p3 == other.p4 and self.p4 == other.p1:
            return True
        if self.p1 == other.p3 and self.p2 == other.p4 and self.p3 == other.p1 and self.p4 == other.p2:
            return True
        if self.p1 == other.p4 and self.p2 == other.p1 and self.p3 == other.p2 and self.p4 == other.p3:
            return True

    @property
    def center(self):
        x = [p.x for p in [self.p1, self.p2, self.p3, self.p4]]
        y = [p.y for p in [self.p1, self.p2, self.p3, self.p4]]
        centroid = Point(sum(x) / 4, sum(y) / 4)
        return centroid

    def draw(self, canvas: ImageDraw.ImageDraw, color='black', width=1, text_color='black'):

        # canvas.rectangle((self.p1.as_tuple, self.p3.as_tuple), outline=color,)
        canvas.line((self.p1.as_tuple, self.p2.as_tuple), color, width)
        canvas.line((self.p2.as_tuple, self.p3.as_tuple), color, width)
        canvas.line((self.p3.as_tuple, self.p4.as_tuple), color, width)
        canvas.line((self.p4.as_tuple, self.p1.as_tuple), color, width)
        if self.text:
            canvas.text((self.p1.x + 3, self.p1.y + 3), self.text, fill=text_color, font=self.font)

    def print_cell(self):
        buffer = ''
        longest = max([len(word) for word in self.text.split("\n")])
        buffer += '┼' + "─" * longest + '┼\n'
        for text_line in self.text.split('\n'):
            buffer += "│" + text_line + ' ' * (longest - len(text_line))
            buffer += "│\n"
        buffer += '┼' + "─" * longest + '┼\n'
        print(buffer)

    def point_inside_polygon(self, point: 'Point', include_edges=True):
        """
        Test if point (x,y) is inside polygon poly.

        poly is N-vertices polygon defined as
        [(x1,y1),...,(xN,yN)] or [(x1,y1),...,(xN,yN),(x1,y1)]
        (function works fine in both cases)

        Geometrical idea: point is inside polygon if horizontal beam
        to the right from point crosses polygon even number of times.
        Works fine for non-convex polygons.
        """
        x, y = point.as_tuple
        x1, y1 = self.p1.as_tuple
        x2, y2 = self.p3.as_tuple
        return x1 < x < x2 and y1 < y < y2


class Table:

    def __init__(self, cells: List[Cell], skeleton: List[List[Cell]], ugly_table: List[List[str]], words, canvas=None):
        self.cells = cells
        self.canvas = canvas
        self.words = words
        self.skeleton = skeleton
        self.ugly_table = ugly_table
        self.global_map = {}
    def to_csv(self, filename):
        """Export table to CSV"""
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Get all rows
            for row_id in sorted(self.global_map.keys()):
                row_data = []
                for col_id in sorted(self.global_map[row_id].keys()):
                    cell = self.global_map[row_id][col_id]
                    row_data.append(cell.text.strip())
                writer.writerow(row_data)
    def to_excel(self, filename):

        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        
        # First, determine the actual table dimensions
        max_row = max(self.global_map.keys()) if self.global_map else 0
        max_col = 0
        for row_data in self.global_map.values():
            if row_data:
                max_col = max(max_col, max(row_data.keys()))
        
        # Track which cells we've already processed
        processed_cells = set()
        
        # First pass: Write all cell values
        for row_id in range(max_row + 1):
            for col_id in range(max_col + 1):
                if row_id in self.global_map and col_id in self.global_map[row_id]:
                    cell = self.global_map[row_id][col_id]
                    # Excel uses 1-based indexing
                    excel_row = row_id + 1
                    excel_col = col_id + 1
                    # Write the cell value
                    ws.cell(row=excel_row, column=excel_col, value=cell.text.strip())
        
        # Second pass: Merge cells
        for row_id in range(max_row + 1):
            for col_id in range(max_col + 1):
                if row_id in self.global_map and col_id in self.global_map[row_id]:
                    cell = self.global_map[row_id][col_id]
                    cell_obj_id = id(cell)
                    
                    # Skip if we've already processed this cell object
                    if cell_obj_id in processed_cells:
                        continue
                    
                    # Find all positions this cell spans
                    positions = []
                    for r_id in range(max_row + 1):
                        for c_id in range(max_col + 1):
                            if (r_id in self.global_map and 
                                c_id in self.global_map[r_id] and 
                                id(self.global_map[r_id][c_id]) == cell_obj_id):
                                positions.append((r_id, c_id))
                    
                    if len(positions) > 1:
                        # Sort positions to find top-left and bottom-right
                        min_row = min(pos[0] for pos in positions)
                        max_row_span = max(pos[0] for pos in positions)
                        min_col = min(pos[1] for pos in positions)
                        max_col_span = max(pos[1] for pos in positions)
                        
                        # Excel uses 1-based indexing
                        start_row = min_row + 1
                        start_col = min_col + 1
                        end_row = max_row_span + 1
                        end_col = max_col_span + 1
                        
                        # Create cell range string
                        start_cell = get_column_letter(start_col) + str(start_row)
                        end_cell = get_column_letter(end_col) + str(end_row)
                        
                        # Merge the cells
                        ws.merge_cells(f'{start_cell}:{end_cell}')
                    
                    # Mark this cell as processed
                    processed_cells.add(cell_obj_id)
        
        # Save the workbook
        wb.save(filename)
        print(f"Saved Excel file with merged cells: {filename}")

    def build_table(self):
        for y, (text_row, skeleton_row) in enumerate(zip(self.ugly_table, self.skeleton)):
            self.global_map[y] = {}
            for x, (text, cell) in enumerate(zip(text_row, skeleton_row)):
                for t_cell in self.cells:
                    if t_cell.point_inside_polygon(cell.center):
                        t_cell.text += text if text else ''
                        self.global_map[y][x] = t_cell

        processed_cells = []
        for cell in tqdm(self.cells, desc='Analyzing cells', unit='cells'):
            if cell in processed_cells:
                continue
            in_words = list(filter(lambda char: cell.point_inside_polygon(
                Point(char['x0'], char['top'])), self.words))
            cell.words = in_words
            processed_cells.append(cell)

        if self.canvas:
            for cell in self.cells:
                # print(cell.get_text())
                cell.draw(self.canvas)

    def get_col(self, col_id) -> List[Cell]:
        col = []
        for row in self.global_map.values():
            col.append(row[col_id])
        return col

    def get_row(self, row_id) -> List[Cell]:
        return list(self.global_map[row_id].values())

    def get_cell(self, x, y) -> Cell:
        return self.global_map[y][x]

    def get_cell_span(self, cell):
        temp = {}
        for row_id, row in self.global_map.items():

            for col_id, t_cell in row.items():
                if t_cell == cell:
                    if not temp.get(row_id, False):
                        temp[row_id] = {}
                    temp[row_id][col_id] = True
        row_span = len(temp)
        col_span = len(list(temp.values())[0])
        return row_span, col_span


class TableExtractor:

    def __init__(self, path):
        self.pdf = pdfplumber.open(path)
        self.draw = False
        self.debug = False

    @staticmethod
    def filter_lines(lines: List[Line]):
        new_lines = []
        lines = list(set(lines))
        la = new_lines.append
        for line1 in tqdm(lines, desc='Filtering lines', unit='lines'):
            if line1 in new_lines:
                continue
            la(line1)
        new_lines = list(set(new_lines))
        return new_lines

    @staticmethod
    def add_skeleton_points(points, line):
        points.append(line.p1)
        points.append(line.p2)

    def build_skeleton(self, lines):
        skeleton_points = []
        skeleton = []
        temp_point = Point(0, 0)
        temp_point.down = temp_point.up = temp_point.left = temp_point.right = True
        vertical = list(filter(lambda l: l.vertical, lines))
        horizontal = list(filter(lambda l: not l.vertical, lines))
        
        for line1 in tqdm(vertical, desc='Building table skeleton', unit='lines'):
            sys.stdout.flush()
            if line1.length < 3.0:
                continue
            self.add_skeleton_points(skeleton_points, line1)
            
            for line2 in horizontal:
                if line1 == line2:
                    continue
                self.add_skeleton_points(skeleton_points, line2)
                
                # Check if intersection exists before creating Point
                intersection = line1.infite_intersect(line2)
                if intersection and intersection[0] is not None and intersection[1] is not None:
                    p1 = Point(intersection)
                    if p1 not in skeleton_points:
                        skeleton_points.append(p1)

                    for n, p in enumerate(skeleton_points):
                        skeleton_points[n].copy(temp_point)
                        if p == p1:
                            p1.copy(p)
                            skeleton_points[n] = p1
        
        skeleton_points = list(set(skeleton_points))
        sorted_y_points = sorted(skeleton_points, key=lambda other: other.y)
        
        for p1 in tqdm(sorted_y_points, desc='Building skeleton cells', unit='point'):
            p2 = p1.get_right(skeleton_points)
            if p2:
                p3 = p2.get_bottom(skeleton_points, right=True)
                p4 = p1.get_bottom(skeleton_points, left=True)
                if p3 and p4:
                    cell = Cell(p1, p2, p3, p4)
                    if cell not in skeleton:
                        skeleton.append(cell)
                    else:
                        continue
        
        return skeleton_points, skeleton

    @staticmethod
    def skeleton_to_2d_table(skeleton: List[Cell]) -> List[List[Cell]]:
        rows = []
        for cell in tqdm(skeleton, desc='Analyzing cell positions', unit='cells'):
            row = tuple(sorted(filter(lambda c: cell.on_same_row(c), skeleton), key=lambda c: c.p1.x))
            rows.append(row)
        rows = list(sorted(list(set(rows)), key=lambda c: c[0].p1.y))
        rows = [list(row) for row in rows]
        return rows

    def parse_page(self, page_n):
        if self.debug:
            print('Parsing page', page_n)
        page = self.pdf.pages[page_n]
        if self.debug:
            print('Rendering page')

        if self.debug:
            print('Finding tables')
        tables = TableFinder(page, {'snap_tolerance': 3, 'join_tolerance': 3})
        if self.debug:
            print('Found', len(tables.tables), 'tables')
        beaut_tables = []
        if self.draw:
            p_im = page.to_image(resolution=100)
            p_im.draw_lines(page.lines)
            p_im.save('page-{}-lines.png'.format(page_n + 1))
        if len(tables.tables) > 5:
            return []
        for n, table in enumerate(tables.tables):
            if self.draw:
                p_im.reset()
                im = Image.new('RGB', (int(page.width), int(page.height)), (255,) * 3)
                canvas = ImageDraw.ImageDraw(im)
            ugly_table = table.extract()
            lines = []  # type: List[Line]
            cells = []  # type: List[Cell]
            for cell in tqdm(table.cells, desc='Parsing cells', unit='cells'):
                # p_im.draw_rect(cell)
                x1, y1, x2, y2 = cell
                p1 = Point(x1, y1)
                p1.right = True
                p1.down = True
                p2 = Point(x2, y1)
                p2.left = True
                p2.down = True
                p3 = Point(x2, y2)
                p3.up = True
                p3.left = True
                p4 = Point(x1, y2)
                p4.up = True
                p4.right = True
                line1 = Line(p1, p2)
                line2 = Line(p2, p3)
                line3 = Line(p3, p4)
                line4 = Line(p4, p1)
                lines.append(line1)
                lines.append(line2)
                lines.append(line3)
                lines.append(line4)
                cell = Cell(p1, p2, p3, p4)
                cells.append(cell)

            # for line in lines:
            #     p_im.draw_line(line.as_tuple)
            lines = self.filter_lines(lines)
            # for line in lines:
            #     line.draw(canvas, color='green')
            if self.draw:
                p_im.save('page-{}-{}_im.png'.format(page_n + 1, n))
                im.save('page-{}-{}.png'.format(page_n + 1, n))
            skeleton_points, skeleton = self.build_skeleton(lines.copy())
            if not skeleton_points:
                continue
            skeleton = self.skeleton_to_2d_table(skeleton)

            # for p in points:
            #     p.draw(canvas)

            beaut_table = Table(cells, skeleton, ugly_table, page.extract_words())
            beaut_table.build_table()
            if self.draw:
                for cell in beaut_table.cells:
                    cell.draw(canvas)
            if self.debug:
                print('Saving rendered table')
            if self.draw:
                p_im.save('page-{}-{}_im.png'.format(page_n + 1, n))
                im.save('page-{}-{}.png'.format(page_n + 1, n))
            if self.draw:
                canvas.rectangle((0,0,page.width,page.height),fill='white') #cleaning canvas
                for row_id, row in enumerate(skeleton):
                    for cell_id, cell in enumerate(row):
                        cell.text = '{}-{}'.format(row_id, cell_id)
                        cell.draw(canvas, color='green',text_color='red')
                im.save('page-{}-{}-skeleton.png'.format(page_n + 1, n))
            beaut_tables.append(beaut_table)

        return beaut_tables


# def pdfplumber_table_to_table():




def sanitize_filename(filename):
    """
    Clean filename by removing/replacing invalid characters, parentheses,
    and replacing spaces with underscores.
    
    Args:
        filename (str): Original filename
    
    Returns:
        str: Sanitized filename safe for all OS
    """
    # Remove file extension if present
    if filename.endswith('.csv'):
        filename = filename[:-4]
    
    # Replace or remove problematic characters
    replacements = {
        '/': '_',
        '\\': '_',
        ':': '-',
        '*': '',
        '?': '',
        '"': '',
        '<': '',
        '>': '',
        '|': '_',
        '(': '',   # remove opening parenthesis
        ')': ''    # remove closing parenthesis
    }
    
    for old, new in replacements.items():
        filename = filename.replace(old, new)
    
    # Remove leading/trailing spaces and dots
    filename = filename.strip('. ')
    
    # Replace all whitespace (spaces, tabs, etc.) with underscores
    filename = re.sub(r'\s+', '_', filename)
    
    # Limit length (max 255 chars for most filesystems)
    if len(filename) > 200:
        filename = filename[:200]
    
    return filename

def extract_table_title(bbox, page, margin=25):
    """
    Extract table title from the area above a table bounding box.
    Looks for patterns like "Table 19. Title text" or "Table 19: Title text"
    Handles multi-line titles and spacing issues in "Table".
    
    Args:
        bbox: Tuple of (x0, y0, x1, y1) bounding box coordinates
        page: pdfplumber page object
        margin: Pixels to search above the table (default: 30)
    
    Returns:
        dict: {
            'has_title': bool,
            'full_title': str or None,
            'clean_title': str or None, # Without "Table X." prefix
            'is_continued': bool,
            'table_number': str or None
        }
    """
    x0, y0, x1, y1 = bbox
    
    # Search area above the table
    search_bbox = (
        max(0, x0),
        max(0, y0 - margin),
        x1,
        y0
    )
    
    try:
        cropped = page.within_bbox(search_bbox)
        text = cropped.extract_text()
        print(text)
        
        if not text:
            return {
                'has_title': False,
                'full_title': None,
                'clean_title': None,
                'is_continued': False,
                'table_number': None
            }
        
        # Clean up the text - also remove spaces within "Table" word
        text = text.strip()
        # Fix "T able" or "Ta ble" etc. -> "Table"
        text = re.sub(r'T\s*a\s*b\s*l\s*e', 'Table', text, flags=re.IGNORECASE)
        
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        
        # Look for table title pattern in the lines
        # Patterns: "Table 19. Title" or "Table 19: Title" or "Table 19 - Title"
        # Also allows optional spaces in "Table" and before delimiter
        table_pattern = re.compile(r'Table\s+(\d+)\s*[\.\:\-]\s*(.+)', re.IGNORECASE)
        
        for i, line in enumerate(lines):
            match = table_pattern.search(line)
            if match:
                table_number = match.group(1)
                title_text = match.group(2).strip()
                
                # Collect continuation lines (lines after the match that don't start with "Table")
                full_title_lines = [line]
                for next_line in lines[i+1:]:
                    # Stop if we hit another table title or empty line
                    if re.search(r'T\s*a\s*b\s*l\s*e\s+\d+', next_line, re.IGNORECASE):
                        break
                    # Add the continuation line
                    full_title_lines.append(next_line)
                    title_text += ' ' + next_line.strip()
                
                # Join all lines for full title
                full_title = ' '.join(full_title_lines)
                
                # Check if it's a continuation
                is_continued = '(continued)' in title_text.lower() or 'continued' in title_text.lower()
                
                # Clean the title (remove "(continued)" if present)
                clean_title = re.sub(r'\s*\(continued\)\s*', '', title_text, flags=re.IGNORECASE)
                clean_title = clean_title.strip()
                
                return {
                    'has_title': True,
                    'full_title': full_title,
                    'clean_title': clean_title,
                    'is_continued': is_continued,
                    'table_number': table_number
                }
        
        return {
            'has_title': False,
            'full_title': None,
            'clean_title': None,
            'is_continued': False,
            'table_number': None
        }
    
    except Exception as e:
        return {
            'has_title': False,
            'full_title': None,
            'clean_title': None,
            'is_continued': False,
            'table_number': None,
            'error': str(e)
        }


def merge_tables(table1, table2):
    """
    Merge two tables by concatenating their rows (excluding header from second table).
    
    Args:
        table1: First table object
        table2: Second table object (continuation)
    
    Returns:
        Table: Merged table object
    """
    # Start from row 1 of table2 (skip header)
    next_row_id = max(table1.global_map.keys()) + 1
    
    for row_id in sorted(table2.global_map.keys()):
        if row_id == 0:  # Skip header row of continuation table
            continue
        
        table1.global_map[next_row_id] = table2.global_map[row_id]
        next_row_id += 1
    
    return table1


def extract_all_tables_auto(path, output_directory, start_page=0, end_page=None, debug=False, output_format='csv'):
    """
    Automatically extract all tables from a PDF by detecting table titles.
    Only processes tables with "Table X." titles and handles continuations.
    
    Args:
        path (str): Path to the PDF file
        output_directory (str): Directory where CSV/Excel files will be saved
        start_page (int): First page to process (0-indexed, default: 0)
        end_page (int): Last page to process (0-indexed, default: None = all pages)
        debug (bool): If True, print detailed processing information
        output_format (str): Output format - 'csv', 'excel', or 'both' (default: 'csv')
    
    Returns:
        dict: Summary of extracted tables with status for each page
    """
    
    # Validate output format
    if output_format not in ['csv', 'excel', 'both']:
        raise ValueError("output_format must be 'csv', 'excel', or 'both'")
    
    # Initialize the PDF extractor
    pdf_interpreter = TableExtractor(path)
    
    # Ensure output directory exists
    output_dir = Path(output_directory)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Determine page range
    total_pages = len(pdf_interpreter.pdf.pages)
    if end_page is None:
        end_page = total_pages - 1
    else:
        end_page = min(end_page, total_pages - 1)
    
    # Track extraction results
    results = {
        "success": [],
        "skipped": [],
        "merged": [],
        "errors": [],
        "total_tables": 0,
        "total_pages_processed": 0
    }
    
    # Store last table for continuation detection
    last_table_info = None  # {table_obj, clean_title, filename, page}
    
    print(f"Processing pages {start_page} to {end_page} ({end_page - start_page + 1} pages)")
    
    # Process each page
    for page_index in range(start_page, end_page + 1):
        page_num = page_index + 1  # 1-indexed for display
        
        if debug:
            print(f"\n{'='*60}")
            print(f"Page {page_num} (index {page_index})")
            print(f"{'='*60}")
        
    
        page = pdf_interpreter.pdf.pages[page_index]
        
        # Extract all potential tables on this page
        all_tables = pdf_interpreter.parse_page(page_index)
        
        if not all_tables:
            if debug:
                print(f"  No tables found")
            results["total_pages_processed"] += 1
            continue
        
        if debug:
            print(f"  Found {len(all_tables)} potential table(s)")
        
        # Process each table
        for table_idx, table in enumerate(all_tables):
            # Get table bounding box
            if not table.cells:
                if debug:
                    print(f"  Table {table_idx}: No cells, skipping")
                continue
            
            min_x = min(cell.p1.x for cell in table.cells)
            min_y = min(cell.p1.y for cell in table.cells)
            max_x = max(cell.p3.x for cell in table.cells)
            max_y = max(cell.p3.y for cell in table.cells)
            bbox = (min_x, min_y, max_x, max_y)
            
            # Check for table title above the table
            title_info = extract_table_title(bbox, page)
            
            if not title_info['has_title']:
                if debug:
                    print(f"  Table {table_idx}: No 'Table X.' title found, skipping")
                results["skipped"].append({
                    "page": page_num,
                    "table_index": table_idx,
                    "reason": "No table title found"
                })
                continue
            
            if debug:
                print(f"  Table {table_idx}: '{title_info['full_title']}'")
                if title_info['is_continued']:
                    print(f"    -> CONTINUATION")
            
            # Handle continuation
            if title_info['is_continued']:
                if last_table_info and last_table_info['clean_title'] == title_info['clean_title']:
                    if debug:
                        print(f"    -> Merging with table from page {last_table_info['page']}")
                    
                    # Merge tables
                    last_table_info['table_obj'] = merge_tables(last_table_info['table_obj'], table)
                    
                    # Re-save the merged table in the appropriate format(s)
                    if output_format in ['csv', 'both']:
                        csv_path = Path(last_table_info['csv_filename'])
                        last_table_info['table_obj'].to_csv(str(csv_path))
                    
                    if output_format in ['excel', 'both']:
                        excel_path = Path(last_table_info['excel_filename'])
                        last_table_info['table_obj'].to_excel(str(excel_path))
                    
                    # Track the merge
                    results["merged"].append({
                        "main_page": last_table_info['page'],
                        "continued_on": page_num,
                        "title": title_info['clean_title']
                    })
                    
                    if debug:
                        print(f"    -> Merged and saved")
                else:
                    if debug:
                        print(f"    -> WARNING: Continuation but no matching previous table")
                    results["skipped"].append({
                        "page": page_num,
                        "table_index": table_idx,
                        "reason": "Continuation with no matching previous table"
                    })
                continue
            
            # New table - save it
            sanitized_title = sanitize_filename(title_info['clean_title'])
            
            # Generate filenames based on output format
            csv_filename = f"{sanitized_title}.csv"
            excel_filename = f"{sanitized_title}.xlsx"
            
            csv_path = output_dir / csv_filename
            excel_path = output_dir / excel_filename
            
            # Save table in the requested format(s)
            saved_files = []
            
            if output_format in ['csv', 'both']:
                table.to_csv(str(csv_path))
                saved_files.append(csv_filename)
                if debug:
                    print(f"    -> Saved CSV: {csv_filename}")
            
            if output_format in ['excel', 'both']:
                table.to_excel(str(excel_path))
                saved_files.append(excel_filename)
                if debug:
                    print(f"    -> Saved Excel: {excel_filename}")
            
            # Store for potential continuation
            last_table_info = {
                'table_obj': table,
                'clean_title': title_info['clean_title'],
                'csv_filename': str(csv_path),
                'excel_filename': str(excel_path),
                'page': page_num
            }
            
            # Track success
            results["success"].append({
                "page": page_num,
                "table_number": title_info['table_number'],
                "title": title_info['clean_title'],
                "files": saved_files,
                "csv_path": str(csv_path) if output_format in ['csv', 'both'] else None,
                "excel_path": str(excel_path) if output_format in ['excel', 'both'] else None
            })
            
            results["total_tables"] += 1
        
        results["total_pages_processed"] += 1
    
    return results